#!/usr/bin/env python
# -*- coding: utf-8 -*-
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from validate_email import validate_email
import openpyxl


def send_email(to_mail):

 
    sender = '...' # сюда логин 
    
    passwd = "..." # сюда пароль
    

    server = smtplib.SMTP('smtp.yandex.ru', 587)
    server.starttls()
    try:
        with open("email3.html", encoding="utf-8") as file:
            template = file.read()
    except IOError:

        return "файл с разметкой письма не найден"

    try:
        server.login(sender, passwd)
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = to_mail
        msg['Subject'] = 'Тема письма'
        msg.attach(MIMEText(template, "html", "utf-8"))
        server.sendmail(sender, to_mail, msg.as_string())

        return "Письмо отправлено"
    except Exception as _ex:
        with open(r"{}\error_send_{}.txt".format(os.getcwd(), to_mail), "w") as file:
            file.write('Email: ' + to_mail + f' Тело ошибки: {_ex}\n')
        return f"{_ex}\n Ошибка отправки"


def validate():
    wookbook = openpyxl.load_workbook("baza3000.xlsx")
    worksheet = wookbook.active
    count_valid = 0
    list_bad_email = []
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if validate_email(col[i].value):
                print(col[i].value)
                print(send_email(str(col[i].value)))
                count_valid += 1
            else:
                list_bad_email.append(col[i].value)
        print('')
        with open(r"{}\list_bad_email.txt".format(os.getcwd()), "w") as file:
            for email in list_bad_email:
                file.write(email + '\n')
        if count_valid == 50:
            print(count_valid)
            return '100 писем'
    print('Итого:', count_valid)


def main():
    validate()
    # print(send_email('itkinov6@gmail.com'))

if __name__ == "__main__":
    main()
