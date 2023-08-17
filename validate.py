import os
from validate_email import validate_email
import openpyxl


def validate():
    wookbook = openpyxl.load_workbook("baza.xlsx")
    worksheet = wookbook.active
    count_valid = 0
    list_bad_email = []
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if validate_email(col[i].value):
                # ('Валидный email', col[i].value, end="\t\t)
                count_valid += 1
            else:

                list_bad_email.append(col[i].value)
        print('')
        with open(r"{}\list_bad_email.txt".format(os.getcwd()), "w") as file:
            for email in list_bad_email:
                file.write(email + '\n')
    print('Итого:', count_valid)


